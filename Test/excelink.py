from openpyxl import Workbook

wb = Workbook()


first_sheet = wb.create_sheet(title='link')

second_sheet =wb.create_sheet("CopyLine")

first_sheet['A1'] = "test_filename.png"

#link_from = first_sheet['A1']
#link_to = ("C:/KerryDemo/TestReport/2017-07-26_183531/TC3/Steps/TC3_DataSet_2_Step2.png")

#link_from.hyperlink = link_to



#second_sheet['A1'] = first_sheet['A1']

second_sheet['A1'] = "test_copy_filename.png"

wb.save(r"./workbook2.xlsx")
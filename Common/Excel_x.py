from Common.CONST import CONST
from pathlib import Path
import openpyxl
import shutil
import random

import time


class Excel():
    def __init__(self, path=CONST.EXCELPATH):
        try:
            self.excelwb = openpyxl.load_workbook(path)
            self.excelst = ''
            self.filepath = path
        except Exception as msg:
            print(msg)

    def Select_Sheet_By_Name(self, name):
        self.excelst = self.excelwb.get_sheet_by_name(name)
        return self.excelst

    def Get_Row_Numbers(self):

        rows = self.excelst.max_row

        for i in range (1, rows + 1):
            if self.excelst.cell(row=i,column=1).value is not None:
                if i != rows:
                    continue
                else:
                    return i
            else:
                return i

    def Get_All_Values_By_ColName(self, colname):

        values = []
        for c in range(1, self.excelst.max_column + 1):
            if self.excelst.cell(row=1, column=c).value == colname:
                for r in range(2, self.Get_Row_Numbers() + 1):
                    if self.excelst.cell(row=r, column=c).value is not None:
                        values.append(self.excelst.cell(row=r, column=c).value)
        return values

    def Get_Value_By_ColName(self, colname, row, path = ''):

        for c in range(1, self.excelst.max_column + 1):
            if self.excelst.cell(row=1, column=c).value == colname:
                if row <= self.excelst.max_row:
                    value = self.excelst.cell(row=row + 1, column=c).value

                    if str(value).find("random") != -1 and path !="":
                        excelTemp = Excel(CONST.EXCELPATH)
                        #excelTemp = Excel(r"..\TestCaseData\Kerry_data_2c.xlsx")
                        excelTemp.Select_Sheet_By_Name("data")
                        randomlist = excelTemp.Get_All_Values_By_ColName(value)
                        print(randomlist)
                        randomvalue = random.choice(randomlist)
                        file = Path(path) / Path("%d_%s_%s.random" % (row, colname, str(randomvalue)))
                        file.touch()
                        #print(str(file))
                        return randomvalue
                    else:
                        return value
                else:
                    raise Exception("Given row number is larger than the rows of the sheet")
        else:
            raise Exception("Cannot find any value")

    def Set_Value_By_ColName(self, content, colname, row, sheetpassed=''):
        if sheetpassed != '':
            sheet = sheetpassed
        else:
            sheet = self.excelst

            for c in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=c).value == colname:
                    if row <= sheet.max_row:
                        sheet.cell(row=row + 1, column=c).value = content
                        break
                    else:
                        raise Exception("Given row number is larger than the rows of the sheet")
            else:
                raise Exception("Cannot find the colname")

    def Set_Sheet_Name(self,old_name,new_name):

        self.excelwb.get_sheet_by_name(old_name).title = new_name

    def Get_Excution_DataSet(self,colname,dataset_name="Data set",keyword="Not Yet"):
        DaList = []
        for c in range(1, self.excelst.max_column + 1):
            if self.excelst.cell(row=1, column=c).value == colname:
                for r in range(1,self.excelst.max_row + 1):
                    if self.excelst.cell(row=r + 1, column=c).value == keyword:
                        if self.Get_Value_By_ColName(dataset_name,r) != '':
                            #print("The value of cell(%d,%d) is %r" % (r,c, self.Get_Value_By_ColName(dataset_name,r)))
                            DaList.append(int(self.Get_Value_By_ColName(dataset_name,r)))
        return DaList

    def Save_Excel(self,path=''):
        if path == '':
            savepath = self.filepath
        else:
            savepath = path

        self.excelwb.save(savepath)
